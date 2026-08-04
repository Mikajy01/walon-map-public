"""Lecture du gabarit Excel et écriture du fichier de sortie.

Le fichier de sortie est produit en travaillant directement sur une copie du
classeur gabarit (`Entête walonmap (avec colonnes en plus).xlsx`) : les
styles, largeurs de colonnes, lignes d'en-tête fusionnées, etc. sont donc
conservés sans logique de copie de style manuelle.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import get_logger

_logger = get_logger("services.excel_service")

# Structure du gabarit, constatée à l'analyse (Étape 1) :
#   ligne 1 = catégorie fusionnée (ZONE, PATRIMOINE, ...), absente pour A-F
#   ligne 2 = en-tête réel de chaque colonne
#   ligne 3+ = données
HEADER_ROW = 2
FIRST_DATA_ROW = 3

# Colonnes contenant des identifiants qui RESSEMBLENT à des nombres (code
# postal, numéro de maison) mais ne le sont pas toujours (numéro de maison
# avec lettre comme "12A", ou "/" quand la valeur est absente) : écrites en
# nombre Excel réel quand la valeur est un entier positif propre (pas de
# zéro de tête, qui serait perdu par une conversion en nombre), sinon en
# texte tel quel. Contrairement à un format Texte forcé sur toute la
# colonne, un vrai nombre ne déclenche jamais l'avertissement Excel
# "nombre stocké en tant que texte" — rien à supprimer, la cause disparaît.
NUMERIC_WHEN_POSSIBLE_COLUMNS = {"B", "E"}


def _valeur_pour_cellule(valeur: str):
    """Convertit en entier si c'est un nombre positif propre (sans zéro de
    tête, qui serait sinon perdu — ex: garder "007" en texte, pas 7),
    sinon renvoie la chaîne telle quelle (ex: "12A", "/")."""
    if valeur.isdigit() and (valeur == "0" or not valeur.startswith("0")):
        return int(valeur)
    return valeur


class ExcelService:
    def __init__(self, template_path: Path) -> None:
        self._template_path = template_path

    def load_output_workbook(self) -> Workbook:
        """Charge une copie du gabarit, prête à être complétée."""
        return openpyxl.load_workbook(self._template_path)

    def get_active_sheet(self, wb: Workbook) -> Worksheet:
        return wb.active

    def get_column_headers(self, ws: Worksheet) -> Dict[str, str]:
        """{lettre_colonne: texte d'en-tête} d'après la ligne HEADER_ROW."""
        headers: Dict[str, str] = {}
        for cell in ws[HEADER_ROW]:
            if cell.value is not None:
                headers[cell.column_letter] = str(cell.value).strip()
        return headers

    def write_parcelles(
        self, ws: Worksheet, valeurs_par_ligne: List[Dict[str, str]], column_order: List[str]
    ) -> None:
        """Écrit une ligne par élément de `valeurs_par_ligne` (chacun étant
        un dict {lettre_colonne: valeur}, typiquement `Parcelle.valeurs` ou
        une entrée de `ProgressStore.all_for_commune`)."""
        for offset, valeurs in enumerate(valeurs_par_ligne):
            row_idx = FIRST_DATA_ROW + offset
            for col_letter in column_order:
                cell = ws[f"{col_letter}{row_idx}"]
                brute = valeurs.get(col_letter, "")
                if col_letter in NUMERIC_WHEN_POSSIBLE_COLUMNS:
                    cell.value = _valeur_pour_cellule(brute)
                    if isinstance(cell.value, int):
                        cell.number_format = "0"
                else:
                    cell.value = brute
        _logger.info("%d ligne(s) écrite(s) à partir de la ligne %d", len(valeurs_par_ligne), FIRST_DATA_ROW)

    def lire_donnees_existantes(self, ws: Worksheet, column_order: List[str]) -> List[Dict[str, str]]:
        """Relit les données d'un Excel déjà rempli — inverse de
        `write_parcelles`, utilisé pour importer un fichier corrigé à la
        main (voir services/sync_service.py). S'arrête à la première ligne
        vide (colonne A), comme `write_parcelles` les produit sans trou.

        Les valeurs numériques (colonnes NUMERIC_WHEN_POSSIBLE_COLUMNS,
        stockées en nombre Excel réel par `write_parcelles`) sont
        reconverties en `str` pour rester comparables telles quelles aux
        valeurs de `ProgressStore` (qui stocke tout en texte, JSON)."""
        lignes: List[Dict[str, str]] = []
        row_idx = FIRST_DATA_ROW
        while True:
            premiere_valeur = ws[f"{column_order[0]}{row_idx}"].value
            if premiere_valeur is None or str(premiere_valeur).strip() == "":
                break
            valeurs: Dict[str, str] = {}
            for col in column_order:
                brute = ws[f"{col}{row_idx}"].value
                valeurs[col] = "" if brute is None else str(brute)
            lignes.append(valeurs)
            row_idx += 1
        _logger.info("%d ligne(s) relue(s) depuis un fichier Excel existant.", len(lignes))
        return lignes

    def save(self, wb: Workbook, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        _logger.info("Fichier de sortie enregistré : %s", output_path)
