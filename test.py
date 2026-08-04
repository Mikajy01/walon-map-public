"""Version d'essai de main.py : ne résout que quelques NOUVELLES parcelles
(2 par défaut) puis affiche le résultat en console, pour valider rapidement
le comportement du programme sans attendre le traitement d'une commune
entière. Utilise exactement la même reprise incrémentale que main.py (même
fichier de sortie `output/<commune>.xlsx`, mêmes parcelles déjà traitées
réutilisées sans être recalculées) — ce script n'est qu'un raccourci pour
lancer un petit lot et l'afficher immédiatement.

Usage :
    python test.py --commune Crisnée
    python test.py --commune Crisnée --limit 5 --debug
"""

from __future__ import annotations

import argparse
import sys

import config
from main import traiter_commune
from services.cache_service import CacheService, ProgressStore
from services.cadastre_service import CadastreService
from services.excel_service import ExcelService
from services.geoportail_service import ArcGISRestClient
from services.layers_service import LayersService
from utils.logger import get_logger, setup_logging
from utils.rate_limiter import RateLimiter

_logger = get_logger("test")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Essai rapide du pipeline (quelques parcelles seulement).")
    parser.add_argument("--commune", required=True, help="Commune à tester.")
    parser.add_argument("--pays", default="Belgique")
    parser.add_argument(
        "--code-postal", action="append", dest="codes_postaux",
        help="Limite au(x) code(s) postal/postaux indiqué(s) (répéter l'option pour en spécifier plusieurs).",
    )
    parser.add_argument("--limit", type=int, default=2, help="Nombre de parcelles à traiter (défaut : 2).")
    parser.add_argument("--debug", action="store_true", help="Active le mode DEBUG (journalisation détaillée).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config.DEBUG = args.debug
    setup_logging(config.LOGS_DIR, debug=config.DEBUG)

    cache_service = CacheService(config.CACHE_DIR)
    progress_store = ProgressStore(config.CACHE_DIR)
    rate_limiter = RateLimiter(config.MAX_REQUESTS_PER_SECOND)
    client = ArcGISRestClient(
        cache=cache_service, rate_limiter=rate_limiter, timeout=config.HTTP_TIMEOUT_SECONDS,
    )
    cadastre_service = CadastreService(client)
    layers_service = LayersService(client, config.LIENS_COMMUNAUX_PATH)
    excel_service = ExcelService(config.TEMPLATE_PATH)

    output_path = traiter_commune(
        args.commune, args.pays,
        cadastre_service, layers_service, progress_store, excel_service,
        limit=args.limit, codes_postaux=args.codes_postaux,
    )

    print(f"\nFichier généré : {output_path}\n")

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    headers = [c.value for c in ws[2]]  # ligne 2 = en-têtes réels
    for row_idx in range(3, 3 + args.limit):
        row = ws[row_idx]
        if row[0].value is None:
            break
        print(f"--- Ligne {row_idx - 2} ---")
        for header, cell in zip(headers, row):
            if header is None:
                continue
            print(f"  {header} : {cell.value}")
        print()

    return 0


if __name__ == "__main__":
    sys.exit(main())
